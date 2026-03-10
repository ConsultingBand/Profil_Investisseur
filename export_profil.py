"""
export_profil.py
Fonctions d'export Excel et PDF pour l'application Profil Investisseur.

Usage dans app.py :
    from export_profil import export_xlsx, export_pdf
"""

import io
from datetime import datetime

# ── EXCEL ─────────────────────────────────────────────────────────────────────

def export_xlsx(report: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    # ── Palette ──
    BLUE       = "2563EB"
    BLUE_DARK  = "1D4ED8"
    BLUE_LT    = "DBEAFE"
    BLUE_BG    = "EEF2FF"
    SLATE      = "0F172A"
    GRAY       = "475569"
    GRAY_LT    = "F8FAFF"
    WHITE      = "FFFFFF"
    BORDER_C   = "CBD5E1"

    # ── Helpers styles ──
    def hfont(size=11, bold=False, color=SLATE, italic=False):
        return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

    def hfill(color):
        return PatternFill("solid", fgColor=color)

    def thin_border(sides="all"):
        s = Side(style="thin", color=BORDER_C)
        n = Side(style=None)
        if sides == "all":
            return Border(left=s, right=s, top=s, bottom=s)
        if sides == "bottom":
            return Border(bottom=s)
        if sides == "top_bottom":
            return Border(top=s, bottom=s)
        return Border()

    def center(wrap=False):
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def left(wrap=False):
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # FEUILLE 1 — SYNTHÈSE
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Synthèse"
    ws1.sheet_view.showGridLines = False

    # Largeurs de colonnes
    for col, width in zip("ABCDEFG", [3, 28, 22, 16, 16, 16, 3]):
        ws1.column_dimensions[col].width = width
    ws1.row_dimensions[1].height = 6

    # ── Bandeau header ──
    ws1.merge_cells("B2:F2")
    c = ws1["B2"]
    c.value = "PROFIL INVESTISSEUR"
    c.font = hfont(9, bold=True, color=BLUE)
    c.fill = hfill(BLUE_BG)
    c.alignment = left()
    ws1.row_dimensions[2].height = 22

    ws1.merge_cells("B3:F3")
    c = ws1["B3"]
    m = report["meta"]
    c.value = f"{m['prenom']} {m['nom']}  ·  {m['date']} à {m['heure']}"
    c.font = hfont(16, bold=True, color=SLATE)
    c.alignment = left()
    ws1.row_dimensions[3].height = 32

    ws1.merge_cells("B4:F4")
    c = ws1["B4"]
    c.value = "Questionnaire de profil investisseur — Document confidentiel"
    c.font = hfont(9, italic=True, color=GRAY)
    c.alignment = left()
    ws1.row_dimensions[4].height = 18

    # Ligne de séparation
    for col in "BCDEF":
        ws1[f"{col}5"].border = Border(bottom=Side(style="medium", color=BLUE))
    ws1.row_dimensions[5].height = 6

    # ── Tableau scores ──
    ws1.row_dimensions[6].height = 10

    SECTIONS = {
        "Connaissances": {"icon": "📘", "max": 30},
        "Risque":        {"icon": "⚖️",  "max": 20},
        "ESG":           {"icon": "🌿", "max": 9},
    }

    # En-têtes colonnes scores
    row = 7
    ws1.row_dimensions[row].height = 22
    for col, label in zip(["B", "C", "D", "E", "F"],
                           ["Dimension", "Profil", "Score", "Score max", "Niveau (%)"]):
        c = ws1[f"{col}{row}"]
        c.value = label
        c.font = hfont(9, bold=True, color=WHITE)
        c.fill = hfill(BLUE)
        c.alignment = center()
        c.border = thin_border()

    # Lignes de scores
    for i, (section, meta) in enumerate(SECTIONS.items()):
        row = 8 + i
        ws1.row_dimensions[row].height = 28
        p = report["profils"][section]
        score = p["score"]
        score_max = p["score_max"]
        bg = BLUE_BG if i % 2 == 0 else GRAY_LT

        vals = [
            ("B", f"{meta['icon']}  {section}", left()),
            ("C", p["label"],                   center()),
            ("D", score,                         center()),
            ("E", score_max,                     center()),
            ("F", f'=D{row}/E{row}',             center()),
        ]
        for col, val, align in vals:
            c = ws1[f"{col}{row}"]
            c.value = val
            c.fill = hfill(bg)
            c.border = thin_border()
            c.alignment = align
            if col in ("D", "E"):
                c.font = hfont(11, bold=True, color=BLUE_DARK)
            elif col == "F":
                c.number_format = "0%"
                c.font = hfont(10, bold=True, color=BLUE)
            else:
                c.font = hfont(10, color=SLATE)

    # ── Descriptions profils ──
    row = 11
    ws1.row_dimensions[row].height = 10

    row = 12
    ws1.merge_cells(f"B{row}:F{row}")
    c = ws1[f"B{row}"]
    c.value = "ANALYSE DÉTAILLÉE"
    c.font = hfont(8, bold=True, color=BLUE)
    c.alignment = left()
    ws1.row_dimensions[row].height = 20

    for section in SECTIONS:
        row += 1
        ws1.row_dimensions[row].height = 14
        p = report["profils"][section]
        icon = SECTIONS[section]["icon"]

        ws1.merge_cells(f"B{row}:C{row}")
        c = ws1[f"B{row}"]
        c.value = f"{icon}  {section}  —  {p['label']}"
        c.font = hfont(10, bold=True, color=SLATE)
        c.fill = hfill(BLUE_BG)
        c.alignment = left()
        for col in "BCDEF":
            ws1[f"{col}{row}"].fill = hfill(BLUE_BG)

        row += 1
        ws1.row_dimensions[row].height = 42
        ws1.merge_cells(f"B{row}:F{row}")
        c = ws1[f"B{row}"]
        c.value = p["description"]
        c.font = hfont(9, italic=True, color=GRAY)
        c.alignment = left(wrap=True)

    # ── Exclusions ESG ──
    if report.get("exclusions"):
        row += 2
        ws1.merge_cells(f"B{row}:F{row}")
        c = ws1[f"B{row}"]
        c.value = "🌿  Exclusions sectorielles ESG"
        c.font = hfont(10, bold=True, color=SLATE)
        c.fill = hfill(BLUE_BG)
        c.alignment = left()
        ws1.row_dimensions[row].height = 20

        row += 1
        ws1.row_dimensions[row].height = 36
        ws1.merge_cells(f"B{row}:F{row}")
        c = ws1[f"B{row}"]
        c.value = "  ·  ".join(report["exclusions"])
        c.font = hfont(9, color=GRAY)
        c.alignment = left(wrap=True)

    # ── Disclaimer ──
    row += 2
    ws1.merge_cells(f"B{row}:F{row}")
    c = ws1[f"B{row}"]
    c.value = (f"Document généré le {m['date']} à {m['heure']}. "
               "Ce document ne constitue pas un conseil en investissement. Confidentiel.")
    c.font = hfont(8, italic=True, color="94A3B8")
    c.alignment = left()

    # ═══════════════════════════════════════════════════════════════════════════
    # FEUILLE 2 — DÉTAIL RÉPONSES
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Détail réponses")
    ws2.sheet_view.showGridLines = False

    for col, width in zip("ABCDEFGH", [3, 8, 12, 52, 12, 10, 10, 3]):
        ws2.column_dimensions[col].width = width
    ws2.row_dimensions[1].height = 6

    # Header
    ws2.merge_cells("B2:G2")
    c = ws2["B2"]
    c.value = f"Détail des réponses  ·  {m['prenom']} {m['nom']}  ·  {m['date']}"
    c.font = hfont(13, bold=True, color=SLATE)
    c.alignment = left()
    ws2.row_dimensions[2].height = 28

    # En-têtes tableau
    row = 3
    ws2.row_dimensions[row].height = 22
    headers = ["ID", "Section", "Question", "Format", "Réponse(s)", "Score"]
    cols    = ["B",  "C",       "D",        "E",      "F",          "G"]
    for col, hdr in zip(cols, headers):
        c = ws2[f"{col}{row}"]
        c.value = hdr
        c.font = hfont(9, bold=True, color=WHITE)
        c.fill = hfill(BLUE)
        c.alignment = center()
        c.border = thin_border()

    current_section = None
    for i, entry in enumerate(report["detail"]):
        row += 1
        ws2.row_dimensions[row].height = 30

        # Séparateur de section
        if entry["section"] != current_section:
            current_section = entry["section"]
            icon = SECTIONS[current_section]["icon"]
            for col in cols:
                c = ws2[f"{col}{row}"]
                c.fill = hfill(BLUE_BG)
                c.border = thin_border("bottom")
            ws2.merge_cells(f"B{row}:G{row}")
            c = ws2[f"B{row}"]
            c.value = f"{icon}  {current_section}"
            c.font = hfont(10, bold=True, color=BLUE_DARK)
            c.fill = hfill(BLUE_BG)
            c.alignment = left()
            row += 1
            ws2.row_dimensions[row].height = 30

        bg = WHITE if i % 2 == 0 else GRAY_LT
        vals = [
            ("B", entry["id"],       center(), hfont(8, bold=True, color=BLUE)),
            ("C", entry["section"],  center(), hfont(9, color=GRAY)),
            ("D", entry["question"], left(True), hfont(9, color=SLATE)),
            ("E", entry["format"],   center(), hfont(8, color=GRAY)),
            ("F", entry["reponses"], left(True), hfont(9, italic=True, color=SLATE)),
            ("G", entry["score"],    center(), hfont(10, bold=True, color=BLUE_DARK)),
        ]
        for col, val, align, font in vals:
            c = ws2[f"{col}{row}"]
            c.value = val
            c.font = font
            c.fill = hfill(bg)
            c.alignment = align
            c.border = thin_border()

    # Total scores par section
    row += 2
    ws2.row_dimensions[row].height = 20
    ws2.merge_cells(f"B{row}:F{row}")
    c = ws2[f"B{row}"]
    c.value = "SCORES FINAUX"
    c.font = hfont(9, bold=True, color=WHITE)
    c.fill = hfill(BLUE)
    c.alignment = center()
    ws2[f"G{row}"].fill = hfill(BLUE)

    for section in SECTIONS:
        row += 1
        ws2.row_dimensions[row].height = 22
        p = report["profils"][section]
        icon = SECTIONS[section]["icon"]
        ws2.merge_cells(f"B{row}:F{row}")
        c = ws2[f"B{row}"]
        c.value = f"{icon}  {section}  —  {p['label']}"
        c.font = hfont(10, bold=True, color=SLATE)
        c.fill = hfill(BLUE_BG)
        c.alignment = left()
        c = ws2[f"G{row}"]
        c.value = f"{p['score']} / {p['score_max']}"
        c.font = hfont(10, bold=True, color=BLUE_DARK)
        c.fill = hfill(BLUE_BG)
        c.alignment = center()

    # Sauvegarde en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ── PDF ────────────────────────────────────────────────────────────────────────

def export_pdf(report: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, white
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import Flowable
    from reportlab.pdfgen import canvas as cv

    # ── Palette ──
    BLUE      = HexColor("#2563EB")
    BLUE_DARK = HexColor("#1D4ED8")
    BLUE_LT   = HexColor("#DBEAFE")
    BLUE_BG   = HexColor("#EEF2FF")
    SLATE     = HexColor("#0F172A")
    SLATE2    = HexColor("#1E293B")
    GRAY      = HexColor("#475569")
    GRAY_LT   = HexColor("#94A3B8")
    GRAY_BG   = HexColor("#F8FAFF")
    BORDER    = HexColor("#CBD5E1")

    W, H = A4
    ML, MR, MT, MB = 20*mm, 20*mm, 28*mm, 20*mm
    TW = W - ML - MR

    # ── Styles texte ──
    def S(name, **kw):
        base = dict(fontName="Helvetica", fontSize=9.5, textColor=GRAY,
                    leading=14, spaceAfter=4)
        base.update(kw)
        return ParagraphStyle(name, **base)

    s_label    = S("label", fontName="Helvetica-Bold", fontSize=7,
                   textColor=BLUE, leading=10, spaceAfter=1)
    s_name     = S("name", fontName="Helvetica-Bold", fontSize=20,
                   textColor=SLATE, leading=24, spaceAfter=2)
    s_meta     = S("meta", fontSize=8.5, textColor=GRAY_LT, leading=12)
    s_h1       = S("h1", fontName="Helvetica-Bold", fontSize=13,
                   textColor=SLATE, leading=17, spaceBefore=14, spaceAfter=5)
    s_h2       = S("h2", fontName="Helvetica-Bold", fontSize=10,
                   textColor=SLATE2, leading=14, spaceBefore=10, spaceAfter=3)
    s_body     = S("body", leading=14, spaceAfter=5)
    s_body_it  = S("body_it", fontName="Helvetica-Oblique",
                   textColor=GRAY, leading=14, spaceAfter=5)
    s_profil   = S("profil", fontName="Helvetica-Bold", fontSize=10.5,
                   textColor=BLUE_DARK, leading=14, spaceAfter=2)
    s_q        = S("q", fontName="Helvetica-Bold", fontSize=9,
                   textColor=SLATE, leading=13, spaceAfter=2)
    s_ans_sel  = S("ans_sel", fontName="Helvetica-Bold", fontSize=9,
                   textColor=BLUE_DARK, leading=13, leftIndent=12, spaceAfter=1)
    s_ans_oth  = S("ans_oth", fontSize=9, textColor=GRAY_LT,
                   leading=13, leftIndent=12, spaceAfter=1)
    s_footer   = S("footer", fontSize=7.5, textColor=GRAY_LT, leading=10)
    s_excl     = S("excl", fontSize=8.5, textColor=GRAY, leading=13)
    s_caption  = S("caption", fontName="Helvetica-Oblique", fontSize=8,
                   textColor=GRAY_LT, leading=11)

    m = report["meta"]

    SECTIONS_META = {
        "Connaissances": {"icon": "📘"},
        "Risque":        {"icon": "⚖️"},
        "ESG":           {"icon": "🌿"},
    }

    # ── Header / Footer callbacks ──
    def on_page(canvas, doc):
        canvas.saveState()
        pn = doc.page

        # Bande bleue haut
        canvas.setFillColor(BLUE)
        canvas.rect(0, H - 14*mm, W, 14*mm, fill=1, stroke=0)
        canvas.setFillColor(white)
        canvas.setFont("Helvetica-Bold", 7.5)
        canvas.drawString(ML, H - 8.5*mm, "PROFIL INVESTISSEUR")
        canvas.setFont("Helvetica", 7.5)
        canvas.drawRightString(W - MR, H - 8.5*mm,
                               f"{m['prenom']} {m['nom']}  ·  {m['date']}")

        # Footer
        canvas.setFillColor(BORDER)
        canvas.rect(ML, MB - 6*mm, TW, 0.4, fill=1, stroke=0)
        canvas.setFillColor(GRAY_LT)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(ML, MB - 10*mm,
                          "Document confidentiel · Ne constitue pas un conseil en investissement")
        canvas.drawRightString(W - MR, MB - 10*mm, f"Page {pn}")

        canvas.restoreState()

    # ── Flowable : barre score ──
    class ScoreBar(Flowable):
        def __init__(self, label, profil, score, score_max, width=TW):
            super().__init__()
            self.label     = label
            self.profil    = profil
            self.score     = score
            self.score_max = score_max
            self.bw        = width
            self.height    = 38

        def wrap(self, *a): return self.bw, self.height

        def draw(self):
            c   = self.canv
            pct = self.score / self.score_max

            # Fond carte
            c.setFillColor(GRAY_BG)
            c.roundRect(0, 0, self.bw, self.height, 5, fill=1, stroke=0)

            # Label + profil
            c.setFillColor(GRAY)
            c.setFont("Helvetica", 8)
            c.drawString(10, 24, self.label)
            c.setFillColor(SLATE)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(10, 12, self.profil)

            # Barre de progression
            bar_x = self.bw * 0.45
            bar_w = self.bw * 0.52
            bar_h = 7
            bar_y = 15
            # fond
            c.setFillColor(BORDER)
            c.roundRect(bar_x, bar_y, bar_w, bar_h, 3, fill=1, stroke=0)
            # remplissage
            if pct > 0:
                c.setFillColor(BLUE)
                c.roundRect(bar_x, bar_y, bar_w * pct, bar_h, 3, fill=1, stroke=0)
            # pourcentage
            c.setFillColor(GRAY_LT)
            c.setFont("Helvetica", 7.5)
            c.drawRightString(self.bw - 4, bar_y + 1, f"{round(pct*100)}%")

    # ── Flowable : séparateur section ──
    class SectionHeader(Flowable):
        def __init__(self, title, width=TW):
            super().__init__()
            self.title = title
            self.bw    = width
            self.height = 20

        def wrap(self, *a): return self.bw, self.height

        def draw(self):
            c = self.canv
            c.setFillColor(BLUE_BG)
            c.roundRect(0, 0, self.bw, self.height, 4, fill=1, stroke=0)
            c.setFillColor(BLUE)
            c.rect(0, 0, 4, self.height, fill=1, stroke=0)
            c.setFillColor(BLUE_DARK)
            c.setFont("Helvetica-Bold", 9.5)
            c.drawString(12, 6, self.title)

    # ── Construction du story ──
    story = []

    # — Couverture / En-tête identité —
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("PROFIL INVESTISSEUR", s_label))
    story.append(Paragraph(f"{m['prenom']} {m['nom']}", s_name))
    story.append(Paragraph(
        f"{m['date']} à {m['heure']}",
        s_meta))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width=TW, thickness=1.5, color=BLUE, spaceAfter=5))

    # — Scores synthèse —
    story.append(Paragraph("Synthèse du profil", s_h1))

    score_bars = []
    for section, meta in SECTIONS_META.items():
        p = report["profils"][section]
        score_bars.append(ScoreBar(
            label=f"{meta['icon']}  {section}",
            profil=p["label"],
            score=p["score"],
            score_max=p["score_max"],
            width=(TW - 8) / 3,
        ))

    story.append(Table(
        [[score_bars[0], score_bars[1], score_bars[2]]],
        colWidths=[(TW - 8) / 3] * 3,
        style=TableStyle([
            ("LEFTPADDING",  (0,0), (-1,-1), 4),
            ("RIGHTPADDING", (0,0), (-1,-1), 4),
            ("TOPPADDING",   (0,0), (-1,-1), 0),
            ("BOTTOMPADDING",(0,0), (-1,-1), 0),
        ])
    ))

    story.append(Spacer(1, 4*mm))

    # — Descriptions profils —
    story.append(Paragraph("Analyse par dimension", s_h1))

    for section, meta in SECTIONS_META.items():
        p = report["profils"][section]
        block = [
            Paragraph(f"{meta['icon']}  {section}  —  {p['label']}", s_profil),
            Paragraph(p["description"], s_body_it),
        ]
        if section == "ESG" and report.get("exclusions"):
            excl_str = "  ·  ".join(report["exclusions"])
            block.append(Paragraph(f"<b>Exclusions sectorielles :</b> {excl_str}", s_excl))

        story.append(KeepTogether(block))
        story.append(Spacer(1, 2*mm))

    story.append(HRFlowable(width=TW, thickness=0.5, color=BORDER, spaceAfter=4))

    # — Détail questionnaire —
    story.append(Paragraph("Détail du questionnaire", s_h1))

    # On reconstruit les questions avec toutes leurs options
    # depuis report["detail"] enrichi des options complètes
    # Le report ne contient que les réponses choisies — on a besoin des QUESTIONS
    # On passe donc QUESTIONS en paramètre via le wrapper dans app.py
    questions_data = report.get("questions_full", [])

    current_section = None
    for entry in report["detail"]:
        if entry["section"] != current_section:
            current_section = entry["section"]
            story.append(Spacer(1, 2*mm))
            story.append(SectionHeader(
                f"{SECTIONS_META[current_section]['icon']}  {current_section}"
            ))
            story.append(Spacer(1, 3*mm))

        # Question
        story.append(Paragraph(entry["question"], s_q))

        # Options — on affiche toutes les options, réponse sélectionnée en bleu
        q_full = next(
            (q for q in questions_data if q["label"] == entry["question"]), None
        )
        chosen = [r.strip() for r in entry["reponses"].split("|")]

        if q_full:
            for label, _ in q_full["answers"]:
                if label in chosen:
                    story.append(Paragraph(f"▶  {label}", s_ans_sel))
                else:
                    story.append(Paragraph(f"○  {label}", s_ans_oth))
        else:
            # Fallback si questions_full non fourni
            for c in chosen:
                story.append(Paragraph(f"▶  {c}", s_ans_sel))

        story.append(Spacer(1, 2*mm))

    # — Disclaimer —
    story.append(HRFlowable(width=TW, thickness=0.5, color=BORDER, spaceAfter=4))
    story.append(Paragraph(
        f"Document généré le {m['date']} à {m['heure']}. "
        "Ce questionnaire est fourni à titre de démonstration uniquement. "
        "Il ne constitue pas une garantie de conformité auprès d'un régulateur "
        "et ne saurait remplacer une procédure de profilage réglementaire complète. "
        "Confidentiel.",
        s_caption
    ))

    # ── Build ──
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=ML, rightMargin=MR,
        topMargin=MT, bottomMargin=MB,
        title=f"Profil Investisseur — {m['prenom']} {m['nom']}",
        author="Profil Investisseur",
    )
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    buffer.seek(0)
    return buffer.getvalue()
